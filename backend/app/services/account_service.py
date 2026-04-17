"""账号管理 service（Phase 5 模块 D）。

- 厂家 / 中介 预注册 + 编辑
- Excel 批量导入（openpyxl）+ CSV 注入防护
- 工人 / 黑名单 只读查询
- 封禁 / 解封 + audit_log
"""
from __future__ import annotations

import io
import logging
import uuid
from typing import Any

from sqlalchemy.orm import Session

from app.core.exceptions import BusinessException
from app.models import SystemConfig, User
from app.services.admin_log_service import write_admin_log

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------

def _gen_external_userid(role: str) -> str:
    return f"pre_{role}_{uuid.uuid4().hex[:8]}"


def _sanitize_cell(value: Any) -> Any:
    """去除 Excel 公式注入前缀（= + - @）。"""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s  # 在前面加单引号让 Excel 当作文本
    return value


def _load_config_int(db: Session, key: str, default: int) -> int:
    cfg = db.query(SystemConfig).filter(SystemConfig.config_key == key).first()
    if not cfg:
        return default
    try:
        return int(cfg.config_value)
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# 列表查询（分页）
# ---------------------------------------------------------------------------

def _base_query(db: Session, role: str, keyword: str | None, status: str | None):
    query = db.query(User).filter(User.role == role)
    if status:
        query = query.filter(User.status == status)
    if keyword:
        k = f"%{keyword}%"
        query = query.filter(
            (User.external_userid.ilike(k)) |
            (User.display_name.ilike(k)) |
            (User.company.ilike(k)) |
            (User.contact_person.ilike(k)) |
            (User.phone.ilike(k))
        )
    return query


def list_users(
    db: Session,
    role: str,
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    status: str | None = None,
) -> tuple[list[User], int]:
    if role not in ("worker", "factory", "broker"):
        raise BusinessException(40101, "无效角色")
    page = max(1, page)
    size = max(1, min(size, 100))
    query = _base_query(db, role, keyword, status)
    total = query.count()
    rows = query.order_by(User.registered_at.desc()).offset((page - 1) * size).limit(size).all()
    return rows, total


def list_blacklist(db: Session, page: int = 1, size: int = 20, keyword: str | None = None) -> tuple[list[User], int]:
    page = max(1, page)
    size = max(1, min(size, 100))
    query = db.query(User).filter(User.status == "blocked")
    if keyword:
        k = f"%{keyword}%"
        query = query.filter(
            (User.external_userid.ilike(k)) |
            (User.display_name.ilike(k)) |
            (User.company.ilike(k))
        )
    total = query.count()
    rows = query.order_by(User.registered_at.desc()).offset((page - 1) * size).limit(size).all()
    return rows, total


def get_user(db: Session, userid: str) -> User:
    user = db.query(User).filter(User.external_userid == userid).first()
    if not user:
        raise BusinessException(40401, "用户不存在")
    return user


# ---------------------------------------------------------------------------
# 预注册
# ---------------------------------------------------------------------------

def pre_register(db: Session, role: str, payload: dict, operator: str) -> User:
    if role not in ("factory", "broker"):
        raise BusinessException(40101, "仅厂家/中介可预注册")

    external = payload.get("external_userid") or _gen_external_userid(role)
    existing = db.query(User).filter(User.external_userid == external).first()
    if existing:
        raise BusinessException(40904, "external_userid 已存在")

    # 按方案 §13.3：
    # - 厂家默认 can_search_workers=1 / can_search_jobs=0（只能检索工人）
    # - 中介默认双向（两个开关默认 True，可由 payload 覆盖）
    if role == "factory":
        can_search_workers = True if payload.get("can_search_workers") is None else bool(payload.get("can_search_workers"))
        can_search_jobs = bool(payload.get("can_search_jobs"))
    else:  # broker
        can_search_workers = True if payload.get("can_search_workers") is None else bool(payload.get("can_search_workers"))
        can_search_jobs = True if payload.get("can_search_jobs") is None else bool(payload.get("can_search_jobs"))

    user = User(
        external_userid=external,
        role=role,
        display_name=payload.get("display_name"),
        company=payload.get("company"),
        contact_person=payload.get("contact_person"),
        phone=payload.get("phone"),
        can_search_jobs=1 if can_search_jobs else 0,
        can_search_workers=1 if can_search_workers else 0,
        status="active",
    )
    db.add(user)
    try:
        db.flush()
    except Exception as exc:  # IntegrityError 等
        db.rollback()
        raise BusinessException(40904, f"预注册失败：{exc}") from exc

    write_admin_log(
        db,
        target_type="user", target_id=external,
        action="reinstate", operator=operator,
        before=None,
        after={"role": role, "display_name": user.display_name, "company": user.company},
        reason="pre_register",
    )
    return user


def update_user(db: Session, userid: str, payload: dict, operator: str) -> User:
    user = get_user(db, userid)
    before = {
        "display_name": user.display_name,
        "company": user.company,
        "contact_person": user.contact_person,
        "phone": user.phone,
        "can_search_jobs": bool(user.can_search_jobs),
        "can_search_workers": bool(user.can_search_workers),
        "external_userid": user.external_userid,
    }

    new_external = payload.get("external_userid")
    if new_external and new_external != user.external_userid:
        # MySQL 不允许直接改主键；一期简化：拒绝改动 external_userid
        raise BusinessException(40101, "external_userid 不可修改（如需变更请联系运维）")

    for k in ("display_name", "company", "contact_person", "phone"):
        if k in payload and payload[k] is not None:
            setattr(user, k, payload[k])
    if "can_search_jobs" in payload and payload["can_search_jobs"] is not None:
        user.can_search_jobs = 1 if payload["can_search_jobs"] else 0
    if "can_search_workers" in payload and payload["can_search_workers"] is not None:
        user.can_search_workers = 1 if payload["can_search_workers"] else 0

    after = {
        "display_name": user.display_name,
        "company": user.company,
        "contact_person": user.contact_person,
        "phone": user.phone,
        "can_search_jobs": bool(user.can_search_jobs),
        "can_search_workers": bool(user.can_search_workers),
    }
    write_admin_log(
        db,
        target_type="user", target_id=user.external_userid,
        action="manual_edit", operator=operator,
        before=before, after=after,
    )
    db.commit()
    db.refresh(user)
    return user


# ---------------------------------------------------------------------------
# Excel 批量导入
# ---------------------------------------------------------------------------

_IMPORT_COLUMNS = [
    "role", "display_name", "company", "contact_person", "phone",
    "can_search_jobs", "can_search_workers", "external_userid",
]


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ("1", "true", "yes", "y", "是", "t")


def import_excel(
    db: Session,
    role: str,
    file_bytes: bytes,
    operator: str,
) -> dict:
    """Excel 批量导入。任意一行失败 → 全部回滚。"""
    if role not in ("factory", "broker"):
        raise BusinessException(40101, "仅厂家/中介可批量导入")

    max_rows = _load_config_int(db, "account.import_max_rows", 500)

    try:
        from openpyxl import load_workbook  # lazy import
    except ImportError as exc:
        raise BusinessException(50001, "服务器未安装 openpyxl") from exc

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise BusinessException(40101, f"Excel 解析失败: {exc}") from exc

    sheet = wb.active
    # 首行为表头
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise BusinessException(40101, "Excel 为空")
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    # 必要列（role 可省略，由路由固定；允许模板省略）
    required_cols = {"display_name", "phone"}
    missing = required_cols - set(headers)
    if missing:
        raise BusinessException(40101, f"缺少必要列: {','.join(missing)}")

    failed: list[dict] = []
    success: list[dict] = []

    rows_iter = sheet.iter_rows(min_row=2, values_only=True)
    for idx, row in enumerate(rows_iter, start=2):
        if idx - 1 > max_rows:
            failed.append({"row": idx, "error": f"超过单次最大行数 {max_rows}"})
            break
        if row is None or all(c is None for c in row):
            continue  # 空行跳过

        cell_map: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if h in _IMPORT_COLUMNS and i < len(row):
                cell_map[h] = _sanitize_cell(row[i])

        payload = {
            "display_name": cell_map.get("display_name"),
            "company": cell_map.get("company"),
            "contact_person": cell_map.get("contact_person"),
            "phone": cell_map.get("phone"),
            "can_search_jobs": _coerce_bool(cell_map.get("can_search_jobs")) if role == "broker" else False,
            "can_search_workers": _coerce_bool(cell_map.get("can_search_workers")) if role == "broker" else True,
            "external_userid": cell_map.get("external_userid"),
        }
        # 简单必填校验
        if not payload.get("display_name"):
            failed.append({"row": idx, "error": "display_name 必填"})
            continue
        if not payload.get("phone"):
            failed.append({"row": idx, "error": "phone 必填"})
            continue

        try:
            pre_register(db, role, payload, operator)
            success.append({"row": idx, "external_userid": payload.get("external_userid")})
        except BusinessException as exc:
            failed.append({"row": idx, "error": exc.message})
        except Exception as exc:
            logger.exception("account import row=%s failed", idx)
            failed.append({"row": idx, "error": str(exc)})

    if failed:
        db.rollback()
        return {"success_count": 0, "failed": failed}

    db.commit()
    return {"success_count": len(success), "failed": []}


# ---------------------------------------------------------------------------
# 封禁 / 解封
# ---------------------------------------------------------------------------

def block_user(db: Session, userid: str, reason: str, operator: str) -> None:
    user = get_user(db, userid)
    if user.status == "blocked":
        raise BusinessException(40904, "用户已被封禁")
    before = {"status": user.status, "blocked_reason": user.blocked_reason}
    user.status = "blocked"
    user.blocked_reason = reason
    write_admin_log(
        db,
        target_type="user", target_id=user.external_userid,
        action="manual_reject", operator=operator,
        before=before,
        after={"status": user.status, "blocked_reason": user.blocked_reason},
        reason=reason,
    )
    db.commit()


def unblock_user(db: Session, userid: str, reason: str, operator: str) -> None:
    user = get_user(db, userid)
    if user.status != "blocked":
        raise BusinessException(40904, "用户未被封禁")
    before = {"status": user.status, "blocked_reason": user.blocked_reason}
    user.status = "active"
    user.blocked_reason = None
    write_admin_log(
        db,
        target_type="user", target_id=user.external_userid,
        action="reinstate", operator=operator,
        before=before,
        after={"status": user.status, "blocked_reason": user.blocked_reason},
        reason=reason,
    )
    db.commit()
